import io
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from typing import List, Dict, Any, Optional, Tuple
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func as sql_func
from sqlalchemy.orm import selectinload

from app.crud.crud_attempt import crud_exam_attempt
from app.schemas import question as schemas_question # Alias for clarity
from app.db import models
from app.crud.crud_question import crud_chapter, crud_question # Need CRUD for chapter lookup/create and question create
from app.schemas.user import UserImportRecord

# --- Configuration for Excel Columns ---
# Define column headers expected/generated
HEADER_ROW = [
    "Chapter Name", "Question Type", "Stem", "Score",
    "Option A", "Option B", "Option C", "Option D", "Option E", # Add more if needed (F, G, ...)
    "Answer", "Explanation", "Grading Policy (MC)", "Partial Score % (MC)", "Specified Score (MC)",
    "Match Type (Fill)" # Add more for other grading strategies if needed
]
# Map header names to schema fields (adjust if your schema field names differ)
COLUMN_MAP = {
    "Chapter Name": "chapter_name",
    "Question Type": "question_type",
    "Stem": "stem",
    "Score": "score",
    "Option A": "option_a",
    "Option B": "option_b",
    "Option C": "option_c",
    "Option D": "option_d",
    "Option E": "option_e",
    "Answer": "answer_str", # Use a temporary field for raw answer string
    "Explanation": "explanation",
    "Grading Policy (MC)": "mc_policy",
    "Partial Score % (MC)": "mc_partial_percent",
    "Specified Score (MC)": "mc_specified_score",
    "Match Type (Fill)": "fill_match_type",
}
# Reverse map for export header lookup
FIELD_TO_HEADER = {v: k for k, v in COLUMN_MAP.items()}
OPTION_COLUMNS = ["Option A", "Option B", "Option C", "Option D", "Option E"]
OPTION_IDS = ["A", "B", "C", "D", "E"] # Corresponding IDs

# --- Import Logic ---

async def _get_or_create_chapter(db: AsyncSession, lib_id: int, chapter_name: str, chapter_cache: Dict[str, int]) -> int:
    """Finds existing chapter by name or creates a new one within the library."""
    if chapter_name in chapter_cache:
        return chapter_cache[chapter_name]

    # Check if chapter exists in DB
    stmt = select(models.Chapter.id).filter_by(question_lib_id=lib_id, name=chapter_name)
    result = await db.execute(stmt)
    chapter_id = result.scalar_one_or_none()

    if chapter_id:
        chapter_cache[chapter_name] = chapter_id
        return chapter_id
    else:
        # Create chapter
        new_chapter_schema = schemas_question.ChapterCreate(question_lib_id=lib_id, name=chapter_name, order_index=0) # Default order
        # Use CRUD create - NOTE: This commits individually, might be slow. Consider batching or adjusting CRUD.
        # For simplicity now, we'll use the existing CRUD which commits.
        # A better approach might involve adding chapters to the session and committing once after the loop.
        # However, we need the ID immediately.
        try:
             # We call create directly, it handles commit and refresh
             created_chapter = await crud_chapter.create(db=db, obj_in=new_chapter_schema)
             chapter_id = created_chapter.id
             chapter_cache[chapter_name] = chapter_id
             return chapter_id
        except ValueError: # If chapter name constraint violated (race condition?)
             # Re-fetch in case of race condition
             result = await db.execute(stmt)
             chapter_id = result.scalar_one_or_none()
             if chapter_id:
                 chapter_cache[chapter_name] = chapter_id
                 return chapter_id
             else: # Should not happen if ValueError was due to name conflict
                 raise ValueError(f"Failed to create or find chapter '{chapter_name}'")


def _parse_row_to_import_schema(row_data: Dict[str, Any]) -> schemas_question.QuestionImportRow:
    """Parses raw row dict into the QuestionImportRow schema, handling potential type errors."""
    # Attempt to convert score to float, handle errors
    try:
        if row_data.get('score') is not None:
            row_data['score'] = float(row_data['score'])
        else:
             # Handle missing score - raise error or use default? Raise for now.
             raise ValueError("Score is missing")
    except (ValueError, TypeError):
        raise ValueError(f"Invalid score format: '{row_data.get('score')}'")

    # Validate question type enum
    q_type = row_data.get('question_type')
    try:
        # This ensures the value is a valid member of the enum
        schemas_question.QuestionTypeEnum(q_type)
    except ValueError:
        valid_types = ", ".join([e.value for e in schemas_question.QuestionTypeEnum])
        raise ValueError(f"Invalid Question Type '{q_type}'. Must be one of: {valid_types}")

    # Basic check for required fields (adjust based on QuestionImportRow definition)
    required = ["chapter_name", "question_type", "stem", "score", "answer_str"]
    for req in required:
        if not row_data.get(req):
             raise ValueError(f"Missing required field: {FIELD_TO_HEADER.get(req, req)}")

    # Use Pydantic validation
    return schemas_question.QuestionImportRow(**{COLUMN_MAP[k]: v for k, v in row_data.items() if k in COLUMN_MAP})


def _build_question_create_schema(
    import_row: schemas_question.QuestionImportRow,
    chapter_id: int
) -> schemas_question.QuestionCreate:
    """Transforms the validated import row into the QuestionCreate schema."""
    q_type = import_row.question_type
    options: Optional[List[schemas_question.QuestionOption]] = None
    answer: Any = None
    grading_strategy: Optional[Dict[str, Any]] = None

    # --- Handle Options ---
    if q_type in [schemas_question.QuestionTypeEnum.single_choice, schemas_question.QuestionTypeEnum.multiple_choice]:
        options = []
        for opt_id, col_name in zip(OPTION_IDS, OPTION_COLUMNS):
            option_text = getattr(import_row, COLUMN_MAP[col_name].lower(), None) # Get option text from schema field
            if option_text: # Only add if text exists
                options.append(schemas_question.QuestionOption(id=opt_id, text=str(option_text)))
        if not options or len(options) < 2:
            raise ValueError("Choice questions require at least two non-empty options (A and B).")

    # --- Handle Answer ---
    answer_str = import_row.answer_str.strip()
    if q_type in [schemas_question.QuestionTypeEnum.single_choice, schemas_question.QuestionTypeEnum.multiple_choice]:
        # Expect comma-separated list of option IDs (A, B, C...)
        answer = [a.strip().upper() for a in answer_str.split(',') if a.strip()]
        if not answer:
             raise ValueError("Answer cannot be empty for choice questions.")
        # Further validation happens in QuestionCreate schema validator
    elif q_type == schemas_question.QuestionTypeEnum.fill_in_blank:
        # Expect semicolon-separated list of answers for each blank
        answer = [a.strip() for a in answer_str.split(';') if a.strip()]
        if not answer:
             raise ValueError("Answer cannot be empty for fill-in-the-blank questions.")
    elif q_type == schemas_question.QuestionTypeEnum.short_answer:
        # Expect the model answer string directly, or empty if no model answer
        answer = answer_str if answer_str else None

    # --- Handle Grading Strategy (Example for MC and Fill) ---
    if q_type == schemas_question.QuestionTypeEnum.multiple_choice and import_row.mc_policy:
        grading_strategy = {"policy": import_row.mc_policy}
        if import_row.mc_partial_percent is not None:
            grading_strategy["partial_score_percent"] = float(import_row.mc_partial_percent)
        if import_row.mc_specified_score is not None:
            grading_strategy["specified_score"] = float(import_row.mc_specified_score)
    elif q_type == schemas_question.QuestionTypeEnum.fill_in_blank and import_row.fill_match_type:
        grading_strategy = {"match_type": import_row.fill_match_type}

    # --- Create the final schema ---
    try:
        question_data = schemas_question.QuestionCreate(
            chapter_id=chapter_id,
            question_type=q_type,
            stem=import_row.stem,
            score=import_row.score,
            options=options,
            answer=answer,
            explanation=import_row.explanation,
            grading_strategy=grading_strategy
        )
        # The QuestionCreate validator will perform final checks
        return question_data
    except Exception as e: # Catch Pydantic validation errors
        raise ValueError(f"Data validation failed: {e}")


async def process_import(
    db: AsyncSession, file_content: bytes, lib_id: int, creator_id: int
) -> schemas_question.QuestionImportResult:
    """Reads an Excel file, processes rows, and attempts to import questions."""
    workbook: Workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True) # data_only=True to get values, not formulas
    sheet: Worksheet = workbook.active # Use the first sheet

    imported_count = 0
    skipped_count = 0
    errors: List[Dict[str, Any]] = []
    chapter_cache: Dict[str, int] = {} # Cache chapter names to IDs

    header = [cell.value for cell in sheet[1]]
    # Basic header validation (check if essential columns are present)
    if not all(h in header for h in ["Chapter Name", "Question Type", "Stem", "Score", "Answer"]):
         raise ValueError("Invalid Excel format. Missing required header columns.")

    for row_index in range(2, sheet.max_row + 1):
        row_values = [cell.value for cell in sheet[row_index]]
        # Create dict from header and row values
        raw_row_data = dict(zip(header, row_values))
        # Filter out empty rows
        if not any(raw_row_data.values()):
            continue

        try:
            # 1. Parse and validate raw row data types
            parsed_row = _parse_row_to_import_schema(raw_row_data)

            # 2. Get or Create Chapter ID
            chapter_id = await _get_or_create_chapter(db, lib_id, parsed_row.chapter_name, chapter_cache)

            # 3. Build QuestionCreate Schema (includes final validation)
            question_create_schema = _build_question_create_schema(parsed_row, chapter_id)

            # 4. Add question to session (commit happens outside loop)
            # Use CRUD, but ideally CRUD shouldn't commit immediately for bulk operations
            # For now, we rely on the endpoint to commit after this function returns
            db_obj_data = question_create_schema.model_dump()
            db_obj_data["creator_id"] = creator_id
            db_obj = models.Question(**db_obj_data)
            db.add(db_obj) # Add to session, don't commit here

            imported_count += 1

        except Exception as e:
            skipped_count += 1
            errors.append({"row": row_index, "error": str(e)})
            # Continue to the next row on error

    # Note: Question count update and commit should happen in the endpoint *after* this function succeeds.

    return schemas_question.QuestionImportResult(
        total_rows=sheet.max_row - 1, # Exclude header
        imported_count=imported_count,
        skipped_count=skipped_count,
        errors=errors
    )


# --- Export Logic ---

async def _get_all_questions_for_lib(db: AsyncSession, lib_id: int) -> List[models.Question]:
     """Fetches all questions for a library, joining chapter info."""
     stmt = (
         select(models.Question)
         .join(models.Chapter)
         .options(selectinload(models.Question.chapter)) # Load chapter object
         .filter(models.Chapter.question_lib_id == lib_id)
         .order_by(models.Chapter.order_index, models.Chapter.name, models.Question.id) # Logical order
     )
     result = await db.execute(stmt)
     return result.scalars().all()


def _format_question_for_export(question: models.Question) -> Dict[str, Any]:
    """Formats a Question model object into a dictionary suitable for an Excel row."""
    row_data = {
        "chapter_name": question.chapter.name if question.chapter else "N/A",
        "question_type": question.question_type.value,
        "stem": question.stem,
        "score": question.score,
        "explanation": question.explanation,
        "answer_str": "", # Formatted answer string
        # Initialize option columns
        **{COLUMN_MAP[opt_col]: "" for opt_col in OPTION_COLUMNS},
        # Initialize grading strategy columns
        "mc_policy": None, "mc_partial_percent": None, "mc_specified_score": None,
        "fill_match_type": None,
    }

    q_type = question.question_type

    # --- Format Options ---
    if q_type in [schemas_question.QuestionTypeEnum.single_choice, schemas_question.QuestionTypeEnum.multiple_choice]:
        # Assuming question.options is stored as a list of dicts [{'id': 'A', 'text': '...'}, ...]
        options_dict = {opt['id']: opt['text'] for opt in (question.options or [])}
        for opt_id, col_name in zip(OPTION_IDS, OPTION_COLUMNS):
            if opt_id in options_dict:
                 row_data[COLUMN_MAP[col_name]] = options_dict[opt_id]

    # --- Format Answer ---
    # Assuming question.answer is stored appropriately (list for choices/fill, string for short)
    answer = question.answer
    if q_type in [schemas_question.QuestionTypeEnum.single_choice, schemas_question.QuestionTypeEnum.multiple_choice]:
        # Format list of IDs as comma-separated string
        row_data["answer_str"] = ", ".join(sorted(answer)) if isinstance(answer, list) else ""
    elif q_type == schemas_question.QuestionTypeEnum.fill_in_blank:
        # Format list of blank answers as semicolon-separated string
        row_data["answer_str"] = "; ".join(answer) if isinstance(answer, list) else ""
    elif q_type == schemas_question.QuestionTypeEnum.short_answer:
        # Use the string directly
        row_data["answer_str"] = answer if isinstance(answer, str) else ""

    # --- Format Grading Strategy ---
    strategy = question.grading_strategy
    if isinstance(strategy, dict):
        if q_type == schemas_question.QuestionTypeEnum.multiple_choice:
             row_data["mc_policy"] = strategy.get("policy")
             row_data["mc_partial_percent"] = strategy.get("partial_score_percent")
             row_data["mc_specified_score"] = strategy.get("specified_score")
        elif q_type == schemas_question.QuestionTypeEnum.fill_in_blank:
             row_data["fill_match_type"] = strategy.get("match_type")

    # Map back to header names for the final row dict
    final_row = {}
    for field, header in FIELD_TO_HEADER.items():
        final_row[header] = row_data.get(field)

    return final_row


async def generate_export(db: AsyncSession, lib_id: int) -> bytes:
    """Fetches questions for a library and generates an Excel file content as bytes."""
    questions = await _get_all_questions_for_lib(db, lib_id)

    workbook = Workbook()
    sheet: Worksheet = workbook.active
    sheet.title = "Questions"

    # Write Header
    sheet.append(HEADER_ROW)

    # Write Data Rows
    for question in questions:
        try:
             formatted_row_dict = _format_question_for_export(question)
             # Ensure row values are in the same order as HEADER_ROW
             row_values = [formatted_row_dict.get(header) for header in HEADER_ROW]
             sheet.append(row_values)
        except Exception as e:
            # Log error or add a comment row in Excel about the problematic question
            print(f"Error formatting question ID {question.id} for export: {e}")
            sheet.append([f"Error exporting question ID {question.id}", str(e)] + [""] * (len(HEADER_ROW) - 2))


    # Save to memory
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    return file_stream.read()

# --- Results Export Logic ---

RESULT_EXPORT_HEADER = [
    "Attempt ID", "Exam ID", "Exam Name", "User ID", "Username", "Full Name",
    "Start Time", "Submit Time", "Duration (Seconds)", "Status",
    "Final Score", "Max Possible Score"
    # Add Q1_Score, Q2_Score etc. if implementing per-question scores
]

def _format_attempt_for_export(attempt: models.ExamAttempt, exam_name: str, max_score: Optional[float]) -> Dict[str, Any]:
    """Formats an ExamAttempt model object for Excel export row."""
    user = getattr(attempt, 'user', None) # User should be loaded by CRUD
    duration_seconds = None
    if attempt.start_time and attempt.submit_time:
        duration_seconds = (attempt.submit_time - attempt.start_time).total_seconds()

    return {
        "Attempt ID": attempt.id,
        "Exam ID": attempt.exam_id,
        "Exam Name": exam_name,
        "User ID": attempt.user_id,
        "Username": user.username if user else "N/A",
        "Full Name": user.fullname if user else "N/A",
        "Start Time": attempt.start_time.isoformat() if attempt.start_time else None,
        "Submit Time": attempt.submit_time.isoformat() if attempt.submit_time else None,
        "Duration (Seconds)": int(duration_seconds) if duration_seconds is not None else None,
        "Status": attempt.status.value,
        "Final Score": float(attempt.final_score) if attempt.final_score is not None else None,
        "Max Possible Score": float(max_score) if max_score is not None else None,
    }

async def generate_results_export(db: AsyncSession, exam_id: int) -> bytes:
    """Fetches results for an exam and generates an Excel file content as bytes."""
    exam = await db.get(models.Exam, exam_id)
    if not exam:
        raise ValueError("Exam not found for export.")

    # Fetch all completed attempts for the exam
    attempts = await crud_exam_attempt.get_exam_results_admin(db=db, exam_id=exam_id, limit=10000) # High limit for export

    # Get max possible score (cached or calculated)
    # Using the same logic as statistics calculation for now
    max_score_query = select(sql_func.sum(models.ExamQuestion.score)).where(models.ExamQuestion.exam_id == exam_id)
    max_score_res = await db.execute(max_score_query)
    max_score_possible = max_score_res.scalar_one_or_none()


    workbook = Workbook()
    sheet: Worksheet = workbook.active
    sheet.title = "Exam Results"

    # Write Header
    sheet.append(RESULT_EXPORT_HEADER)

    # Write Data Rows
    for attempt in attempts:
        try:
             formatted_row_dict = _format_attempt_for_export(attempt, exam.name, max_score_possible)
             # Ensure row values are in the same order as HEADER_ROW
             row_values = [formatted_row_dict.get(header) for header in RESULT_EXPORT_HEADER]
             sheet.append(row_values)
        except Exception as e:
            print(f"Error formatting attempt ID {attempt.id} for export: {e}")
            sheet.append([f"Error exporting attempt ID {attempt.id}", str(e)] + [""] * (len(RESULT_EXPORT_HEADER) - 2))

    # Save to memory
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    return file_stream.read()

# --- User Import Logic ---

# Define expected header columns for user import
USER_IMPORT_HEADER = [
    "username", "email", "fullname", "password", "role_names", "group_names"
]
USER_IMPORT_REQUIRED_HEADERS = ["username", "password"] # Minimum required

def parse_user_import_file(file_content: bytes) -> List[UserImportRecord]:
    """
    Parses an Excel file (.xlsx) content for user bulk import.

    Args:
        file_content: Bytes content of the Excel file.

    Returns:
        A list of UserImportRecord objects parsed from the file.

    Raises:
        ValueError: If the header is invalid or data is missing/malformed.
    """
    users: List[UserImportRecord] = []
    try:
        workbook: Workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        sheet: Worksheet = workbook.active

        header_row = [cell.value for cell in sheet[1]] # First row is header

        # Validate header
        header_map: Dict[str, int] = {}
        processed_headers = set()
        for idx, header in enumerate(header_row):
            normalized_header = str(header).strip().lower() if header else None
            if normalized_header in USER_IMPORT_HEADER and normalized_header not in processed_headers:
                header_map[normalized_header] = idx
                processed_headers.add(normalized_header)

        missing_required = [h for h in USER_IMPORT_REQUIRED_HEADERS if h not in header_map]
        if missing_required:
             raise ValueError(f"Missing required header columns: {', '.join(missing_required)}")

        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2): # Start from second row
            row_values = [cell.value for cell in row]
            user_data = {}
            has_data = False
            for header_name, col_idx in header_map.items():
                 if col_idx < len(row_values):
                     value = row_values[col_idx]
                     if value is not None:
                         has_data = True
                         # Handle comma-separated roles/groups
                         if header_name in ["role_names", "group_names"] and isinstance(value, str):
                              user_data[header_name] = [name.strip() for name in value.split(',') if name.strip()]
                         else:
                              user_data[header_name] = value
                     else:
                         # Keep None for optional fields if cell is empty
                         if header_name in ["email", "fullname", "role_names", "group_names"]:
                            user_data[header_name] = None

            if not has_data: # Skip entirely empty rows
                 continue

            # Basic validation (Pydantic will do more)
            if not user_data.get("username") or not user_data.get("password"):
                 raise ValueError(f"Row {row_idx}: Missing required username or password.")

            try:
                 # Use Pydantic for validation and type conversion
                 user_record = UserImportRecord(**user_data)
                 users.append(user_record)
            except Exception as e: # Catch Pydantic validation errors
                 raise ValueError(f"Row {row_idx}: Invalid data format - {e}")

    except ValueError as ve:
         raise ve # Re-raise validation errors
    except Exception as e:
        # Log the exception e
        print(f"Error reading user import file: {e}")
        raise ValueError(f"Failed to read or parse the Excel file. Ensure it's a valid .xlsx file. Error: {e}")

    return users