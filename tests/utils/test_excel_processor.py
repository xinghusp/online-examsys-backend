import json

import pytest
import io
from typing import List, Dict, Any
from unittest.mock import patch, MagicMock, AsyncMock

from openpyxl import Workbook # We might need this to create mock file content
from openpyxl.reader.excel import load_workbook

from app.schemas.question import QuestionTypeEnum
from app.utils import excel_processor
from app.schemas import user as schemas_user, ExamAttemptStatusEnum
from app.schemas import question as schemas_question
from app.schemas import grading as schemas_grading
# Import models if needed for generating export data
# from app import models

# TODO: no completed yet
# Helper function to create mock Excel file bytes
def create_mock_excel_bytes(header: List[str], data_rows: List[List[Any]]) -> bytes:
    """Creates byte content for a simple Excel file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in data_rows:
        sheet.append(row)
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()

# --- Test User Import Parsing ---

VALID_USER_HEADER = ["username", "email", "fullname", "password", "role_names", "group_names"]
VALID_USER_DATA = [
    ["testuser1", "user1@test.com", "Test User One", "pass1", "Student", "Group A"],
    ["testuser2", "user2@test.com", "Test User Two", "pass2", "Student, Teacher", "Group B, Group A"],
    ["testuser3", None, None, "pass3", None, None], # Minimal required
]

def test_parse_user_import_file_success():
    """Test parsing a valid user import Excel file."""
    file_bytes = create_mock_excel_bytes(VALID_USER_HEADER, VALID_USER_DATA)
    parsed_users = excel_processor.parse_user_import_file(file_bytes)

    assert len(parsed_users) == 3
    assert isinstance(parsed_users[0], schemas_user.UserImportRecord)
    assert parsed_users[0].username == "testuser1"
    assert parsed_users[0].email == "user1@test.com"
    assert parsed_users[0].fullname == "Test User One"
    assert parsed_users[0].password == "pass1"
    assert parsed_users[0].role_names == ["Student"]
    assert parsed_users[0].group_names == ["Group A"]

    assert parsed_users[1].username == "testuser2"
    assert parsed_users[1].role_names == ["Student", "Teacher"] # Check comma parsing
    assert parsed_users[1].group_names == ["Group B", "Group A"]

    assert parsed_users[2].username == "testuser3"
    assert parsed_users[2].email is None
    assert parsed_users[2].fullname is None
    assert parsed_users[2].password == "pass3"
    assert parsed_users[2].role_names is None
    assert parsed_users[2].group_names is None

def test_parse_user_import_file_missing_required_header():
    """Test parsing with missing required header columns."""
    header = ["username", "email", "fullname"] # Missing 'password'
    data = [["user1", "email", "name"]]
    file_bytes = create_mock_excel_bytes(header, data)

    with pytest.raises(ValueError, match="Missing required header columns: password"):
        excel_processor.parse_user_import_file(file_bytes)

def test_parse_user_import_file_missing_required_data():
    """Test parsing with a row missing required data (username/password)."""
    header = VALID_USER_HEADER
    data = [
        ["user1", "email", "name", "pass1", None, None],
        [None, "email2", "name2", "pass2", None, None], # Missing username
    ]
    file_bytes = create_mock_excel_bytes(header, data)

    with pytest.raises(ValueError, match="Row 3: Missing required username or password"): # Row 3 because header is 1, data starts at 2
        excel_processor.parse_user_import_file(file_bytes)

def test_parse_user_import_file_invalid_data_type():
    """Test parsing with data that fails Pydantic validation (e.g., bad email)."""
    header = VALID_USER_HEADER
    data = [
        ["user1", "not-an-email", "name", "pass1", None, None],
    ]
    file_bytes = create_mock_excel_bytes(header, data)

    with pytest.raises(ValueError, match="Row 2: Invalid data format"):
        excel_processor.parse_user_import_file(file_bytes)

def test_parse_user_import_file_empty_file():
    """Test parsing an empty or header-only file."""
    header_only_bytes = create_mock_excel_bytes(VALID_USER_HEADER, [])
    parsed_users = excel_processor.parse_user_import_file(header_only_bytes)
    assert len(parsed_users) == 0

    empty_file_bytes = create_mock_excel_bytes([], [])
    # This might raise an error depending on header validation, adjust assertion
    with pytest.raises(ValueError, match="Missing required header columns"):
         excel_processor.parse_user_import_file(empty_file_bytes)

@patch('app.utils.excel_processor.load_workbook')
def test_parse_user_import_file_read_error(mock_load_workbook):
    """Test handling of errors during file reading/loading."""
    mock_load_workbook.side_effect = Exception("Failed to load workbook")
    dummy_bytes = b"dummy excel content"

    with pytest.raises(ValueError, match="Failed to read or parse the Excel file"):
        excel_processor.parse_user_import_file(dummy_bytes)

# --- Test Question Import Parsing ---

# Define expected header columns for question import
QUESTION_IMPORT_HEADER = [
    "content", "question_type", "options", "answers", "score", "explanation"
]
QUESTION_IMPORT_REQUIRED = ["content", "question_type", "score"]

# Sample valid question data
VALID_QUESTION_DATA = [
    ["MCQ Content 1", QuestionTypeEnum.multiple_choice, json.dumps({"A": "Opt A", "B": "Opt B"}), json.dumps(["A"]), 5.0, "Explain 1"],
    ["FIB Content 1 __blank__", QuestionTypeEnum.fill_in_blank, None, json.dumps(["Answer1"]), 3.0, "Explain FIB"],
    ["SAQ Content 1", QuestionTypeEnum.short_answer, None, None, 10.0, "Explain SAQ"], # No predefined answer
    ["MCQ Content 2", QuestionTypeEnum.multiple_choice, '{"A":"Option A","B":"Option B","C":"Option C"}', '["B","C"]', 6.0, None], # Multi-answer MCQ
]

def test_parse_question_import_file_success():
    """Test parsing a valid question import Excel file."""
    file_bytes = create_mock_excel_bytes(QUESTION_IMPORT_HEADER, VALID_QUESTION_DATA)
    parsed_questions = excel_processor.parse_question_import_file(file_bytes)

    assert len(parsed_questions) == 5
    q1 = parsed_questions[0]
    assert isinstance(q1, schemas_question.QuestionImportRecord)
    assert q1.content == "MCQ Content 1"
    assert q1.question_type == schemas_exam.QuestionTypeEnum.multiple_choice
    assert q1.options == {"A": "Opt A", "B": "Opt B"} # Check JSON parsing
    assert q1.answers == ["A"]
    assert q1.score == 5.0
    assert q1.explanation == "Explain 1"

    q2 = parsed_questions[1]
    assert q2.question_type == schemas_exam.QuestionTypeEnum.true_false
    assert q2.options is None
    assert q2.answers == [True]
    assert q2.score == 2.0
    assert q2.explanation is None

    q3 = parsed_questions[2]
    assert q3.question_type == schemas_exam.QuestionTypeEnum.fill_in_blank
    assert q3.answers == ["Answer1"]

    q4 = parsed_questions[3]
    assert q4.question_type == schemas_exam.QuestionTypeEnum.short_answer
    assert q4.answers is None # No answer provided is valid
    assert q4.score == 10.0

    q5 = parsed_questions[4]
    assert q5.question_type == schemas_exam.QuestionTypeEnum.multiple_choice
    assert q5.answers == ["B", "C"] # Check multi-answer JSON parsing

def test_parse_question_import_file_invalid_header():
    """Test parsing with missing required question header columns."""
    header = ["content", "question_type"] # Missing 'score'
    data = [["Content", "multiple_choice"]]
    file_bytes = create_mock_excel_bytes(header, data)

    with pytest.raises(ValueError, match="Missing required header columns: score"):
        excel_processor.parse_question_import_file(file_bytes)

def test_parse_question_import_file_invalid_data_type():
    """Test parsing with invalid data (e.g., non-numeric score, bad JSON)."""
    header = QUESTION_IMPORT_HEADER
    data = [
        ["MCQ Content 1", "multiple_choice", '{"A":"Opt A"}', '["A"]', "five", "Explain 1"], # Invalid score
    ]
    file_bytes = create_mock_excel_bytes(header, data)
    with pytest.raises(ValueError, match="Row 2: Invalid data format.*score"):
        excel_processor.parse_question_import_file(file_bytes)

    data = [
         ["MCQ Content 2", "multiple_choice", '{"A":"Opt A", B:"Opt B"}', '["A"]', 5.0, "Explain 2"], # Invalid JSON (unquoted key)
    ]
    file_bytes = create_mock_excel_bytes(header, data)
    with pytest.raises(ValueError, match="Row 2: Invalid JSON in options"):
         excel_processor.parse_question_import_file(file_bytes)

    data = [
         ["MCQ Content 3", "multiple_choice", '{"A":"Opt A"}', '[A]', 5.0, "Explain 3"], # Invalid JSON array
    ]
    file_bytes = create_mock_excel_bytes(header, data)
    with pytest.raises(ValueError, match="Row 2: Invalid JSON in answers"):
         excel_processor.parse_question_import_file(file_bytes)

def test_parse_question_import_file_missing_required_data():
    """Test parsing with a row missing required question data (content)."""
    header = QUESTION_IMPORT_HEADER
    data = [
        [None, "multiple_choice", '{"A":"Opt A"}', '["A"]', 5.0, "Explain 1"], # Missing content
    ]
    file_bytes = create_mock_excel_bytes(header, data)
    with pytest.raises(ValueError, match="Row 2: Missing required content, question_type, or score"):
        excel_processor.parse_question_import_file(file_bytes)

@patch('app.utils.excel_processor.load_workbook')
def test_parse_question_import_file_read_error(mock_load_workbook):
    """Test handling of errors during question file reading/loading."""
    mock_load_workbook.side_effect = Exception("Failed to load workbook")
    dummy_bytes = b"dummy excel content"

    with pytest.raises(ValueError, match="Failed to read or parse the Excel file"):
        excel_processor.parse_question_import_file(dummy_bytes)


# --- Test Question Export Generation ---

# Mock question data (can be dicts mimicking model attributes)
MOCK_QUESTIONS_FOR_EXPORT = [
    MagicMock(
        id=1, content="MCQ Export 1", question_type=QuestionTypeEnum.multiple_choice,
        options={"A": "Opt A", "B": "Opt B"}, answers=["A"], score=5.0, explanation="Explain 1",
        chapter_id=1, library_id=1
    )
    # Add more mocks as needed
]



# --- Test Results Export Generation ---

# Mock exam and attempt result data
MOCK_EXAM_FOR_EXPORT = MagicMock(
    id=1, name="Test Exam Export", total_score=None # Max score calculated separately
)

MOCK_ATTEMPTS_FOR_EXPORT = [
    schemas_grading.AttemptResultAdmin( # Use the Pydantic schema directly
        id=10, exam_id=1, user_id=3, username="student_user", fullname="Student User",
        status=ExamAttemptStatusEnum.graded,
        final_score=85.5, start_time="2025-04-19T10:00:00Z", submit_time="2025-04-19T11:00:00Z",
        time_taken_seconds=3600, max_score_possible=100.0 # Assume max score is fetched/calculated
    ),
    schemas_grading.AttemptResultAdmin(
        id=11, exam_id=1, user_id=4, username="another_student", fullname="Another Student",
        status=ExamAttemptStatusEnum.graded,
        final_score=70.0, start_time="2025-04-19T10:05:00Z", submit_time="2025-04-19T11:02:00Z",
        time_taken_seconds=3420, max_score_possible=100.0
    ),
     schemas_grading.AttemptResultAdmin(
        id=12, exam_id=1, user_id=5, username="late_student", fullname="Late Student",
        status=ExamAttemptStatusEnum.in_progress, # Test different statuses
        final_score=None, start_time="2025-04-19T10:10:00Z", submit_time=None,
        time_taken_seconds=None, max_score_possible=100.0
    ),
]

# Mock the necessary CRUD/DB calls within the export function
@patch('app.utils.excel_processor.crud_exam_attempt.get_exam_statistics_admin', new_callable=AsyncMock)
@patch('app.utils.excel_processor.crud_exam_attempt.get_exam_results_admin', new_callable=AsyncMock)
@patch('app.utils.excel_processor.db_func.get_db_session') # If using a context managed session
@patch('app.utils.excel_processor.db_func.get_object_or_404') # Or direct db.get mock
@pytest.mark.asyncio
async def test_generate_results_export_success(mock_get_exam, mock_get_db, mock_get_results, mock_get_stats):
    """Test generating a valid results export Excel file."""
    exam_id = 1
    db_mock = AsyncMock() # Mock the session if get_db_session yields it

    # Configure mocks
    mock_get_exam.return_value = MOCK_EXAM_FOR_EXPORT # Mock fetching the exam
    mock_get_db.return_value.__aenter__.return_value = db_mock # If using async with get_db_session():
    mock_get_results.return_value = MOCK_ATTEMPTS_FOR_EXPORT # Mock fetching results
    mock_get_stats.return_value = {"max_score_possible": 100.0} # Mock fetching stats

    # Call the export function
    file_bytes = await excel_processor.generate_results_export(exam_id=exam_id)

    assert file_bytes is not None
    assert len(file_bytes) > 0

    # Optional: Verify content
    workbook = load_workbook(filename=io.BytesIO(file_bytes))
    sheet = workbook.active
    assert sheet.max_row == len(MOCK_ATTEMPTS_FOR_EXPORT) + 1
    assert sheet.cell(row=1, column=1).value == "Attempt ID"
    assert sheet.cell(row=1, column=4).value == "Full Name"
    assert sheet.cell(row=1, column=6).value == "Status"
    assert sheet.cell(row=1, column=7).value == "Score"
    assert sheet.cell(row=1, column=8).value == "Max Score"
    assert sheet.cell(row=1, column=9).value == "Percentage (%)" # Check calculated column

    assert sheet.cell(row=2, column=1).value == MOCK_ATTEMPTS_FOR_EXPORT[0].id
    assert sheet.cell(row=2, column=4).value == MOCK_ATTEMPTS_FOR_EXPORT[0].fullname
    assert sheet.cell(row=2, column=6).value == MOCK_ATTEMPTS_FOR_EXPORT[0].status.value
    assert sheet.cell(row=2, column=7).value == MOCK_ATTEMPTS_FOR_EXPORT[0].final_score
    assert sheet.cell(row=2, column=8).value == 100.0 # Max score from stats
    assert sheet.cell(row=2, column=9).value == 85.5 # Percentage calculation

    assert sheet.cell(row=4, column=6).value == MOCK_ATTEMPTS_FOR_EXPORT[2].status.value # Check in_progress status
    assert sheet.cell(row=4, column=7).value is None # Score should be empty
    assert sheet.cell(row=4, column=9).value == 0.0 # Percentage for no score

@patch('app.utils.excel_processor.db_func.get_object_or_404') # Mock the function that fetches the exam
@pytest.mark.asyncio
async def test_generate_results_export_exam_not_found(mock_get_exam):
    """Test results export when the exam ID is not found."""
    exam_id = 999
    # Configure mock to raise the equivalent of a 404 (e.g., ValueError or specific exception)
    # If get_object_or_404 raises HTTPException, we might need to adjust the expected exception
    # For simplicity, let's assume it raises ValueError if the underlying get fails
    mock_get_exam.side_effect = ValueError(f"Exam with id {exam_id} not found")

    with pytest.raises(ValueError, match=f"Exam with id {exam_id} not found"):
        await excel_processor.generate_results_export(exam_id=exam_id)
